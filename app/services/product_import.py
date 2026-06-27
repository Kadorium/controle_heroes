"""Importação CSV/XLSX de produtos — preview e commit."""

from __future__ import annotations

import csv
import io
from decimal import Decimal
from typing import Any

from sqlalchemy.orm import Session

from app.core.parse import optional_decimal
from app.models import Product
from app.services.product_catalog import LIFECYCLE_ACTIVE

REQUIRED_COLUMNS = {"sku_code", "description", "product_group", "lifecycle_status"}
OPTIONAL_COLUMNS = {
    "ncm",
    "product_subgroup",
    "supplier_code",
    "country_of_origin",
    "unit_of_measure",
    "weight_kg",
    "volume_m3",
    "category",
    "launch_date",
    "fiscal_description",
    "commercial_notes",
}

ALL_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS

# Planilha Heroes (heroes_produtos_epic_import_final.csv)
HEROES_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "sku_code": ("sku_code", "sku_sugerido"),
    "description": ("description", "nome_produto", "descricao_comercial_curta"),
    "product_group": ("product_group", "grupo"),
    "product_subgroup": ("product_subgroup", "subgrupo"),
    "lifecycle_status": ("lifecycle_status", "status_produto"),
    "category": ("category", "categoria_operacional", "tipo_produto"),
    "supplier_code": ("supplier_code", "codigo_fornecedor_ou_heroes"),
    "country_of_origin": ("country_of_origin", "pais_origem"),
    "unit_of_measure": ("unit_of_measure", "unidade_medida"),
    "ncm": ("ncm",),
    "fiscal_description": ("fiscal_description", "descricao_fiscal"),
    "weight_kg": ("weight_kg", "peso_kg"),
    "volume_m3": ("volume_m3",),
    "launch_date": ("launch_date", "data_lancamento", "linha_colecao"),
    "commercial_notes": ("commercial_notes", "observacoes", "descricao_comercial_longa"),
}

LIFECYCLE_ALIASES = {
    "ACTIVE": "ACTIVE",
    "DISCONTINUED": "DISCONTINUED",
    "DRAFT": "DRAFT",
    "ARCHIVED": "ARCHIVED",
    "UNKNOWN": "DRAFT",
}


def _pick(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        val = row.get(key, "")
        if val:
            return val.strip()
    return ""


def _normalize_launch_date(val: str) -> str | None:
    if not val:
        return None
    v = val.strip()
    if len(v) == 4 and v.isdigit():
        return f"{v}-01-01"
    return v


def _canonical_import_row(row: dict[str, str]) -> dict[str, str]:
    out: dict[str, str] = {}
    for canonical, aliases in HEROES_FIELD_ALIASES.items():
        out[canonical] = _pick(row, *aliases)
    status_raw = out.get("lifecycle_status", "").upper()
    out["lifecycle_status"] = LIFECYCLE_ALIASES.get(status_raw, status_raw or LIFECYCLE_ACTIVE)
    out["launch_date"] = _normalize_launch_date(out.get("launch_date", "")) or ""
    country = out.get("country_of_origin", "").upper()
    if country == "ITALIA":
        out["country_of_origin"] = "IT"
    cat = out.get("category", "")
    if cat:
        out["category"] = cat[:32]
    return out


def _normalize_header(name: str) -> str:
    return name.strip().lower().replace(" ", "_")


def _parse_csv_text(content: str) -> list[dict[str, str]]:
    reader = csv.DictReader(io.StringIO(content))
    if not reader.fieldnames:
        return []
    mapping = {_normalize_header(h): h for h in reader.fieldnames if h}
    rows: list[dict[str, str]] = []
    for raw in reader:
        row: dict[str, str] = {}
        for norm, orig in mapping.items():
            val = raw.get(orig, "")
            row[norm] = val.strip() if val is not None else ""
        rows.append(row)
    return rows


def _parse_xlsx_bytes(data: bytes) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return []
    headers = [_normalize_header(str(h or "")) for h in header_row]
    rows: list[dict[str, str]] = []
    for values in rows_iter:
        row: dict[str, str] = {}
        empty = True
        for i, key in enumerate(headers):
            if not key:
                continue
            val = values[i] if i < len(values) else None
            s = "" if val is None else str(val).strip()
            if s:
                empty = False
            row[key] = s
        if not empty:
            rows.append(row)
    return rows


def parse_product_file(content: bytes, *, filename: str = "") -> list[dict[str, str]]:
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        return _parse_xlsx_bytes(content)
    text = content.decode("utf-8-sig", errors="replace")
    return _parse_csv_text(text)


def _validate_row(row: dict[str, str], row_number: int) -> dict[str, Any]:
    errors: list[str] = []
    sku = row.get("sku_code", "")
    desc = row.get("description", "")
    group = row.get("product_group", "")
    status = row.get("lifecycle_status", "") or LIFECYCLE_ACTIVE

    if not sku:
        errors.append("sku_code obrigatório")
    if not desc:
        errors.append("description obrigatório")
    if not group:
        errors.append("product_group obrigatório")
    if status not in ("ACTIVE", "DISCONTINUED", "DRAFT", "ARCHIVED"):
        errors.append("lifecycle_status inválido")

    data: dict[str, Any] | None = None
    if not errors:
        data = {
            "sku_code": sku,
            "description": desc,
            "product_group": group,
            "lifecycle_status": status,
            "ncm": row.get("ncm") or None,
            "product_subgroup": row.get("product_subgroup") or None,
            "supplier_code": row.get("supplier_code") or None,
            "country_of_origin": row.get("country_of_origin") or None,
            "unit_of_measure": row.get("unit_of_measure") or None,
            "weight_kg": optional_decimal(row.get("weight_kg") or None),
            "volume_m3": optional_decimal(row.get("volume_m3") or None),
            "category": row.get("category") or "OTHER",
            "launch_date": row.get("launch_date") or None,
            "fiscal_description": row.get("fiscal_description") or None,
            "commercial_notes": row.get("commercial_notes") or None,
        }

    return {
        "row_number": row_number,
        "sku_code": sku or None,
        "valid": len(errors) == 0,
        "errors": errors,
        "data": data,
    }


def preview_product_import(content: bytes, *, filename: str = "") -> dict:
    raw_rows = parse_product_file(content, filename=filename)
    preview_rows = [_validate_row(_canonical_import_row(r), i + 2) for i, r in enumerate(raw_rows)]
    valid_count = sum(1 for r in preview_rows if r["valid"])
    return {
        "valid_count": valid_count,
        "invalid_count": len(preview_rows) - valid_count,
        "rows": preview_rows,
    }


def commit_product_import(
    db: Session,
    rows: list[dict],
    *,
    user_id: int,
    confirm: bool,
) -> dict:
    if not confirm:
        raise ValueError("confirm=true obrigatório para commit")

    created = updated = skipped = 0
    errors: list[str] = []

    for row in rows:
        data = row.get("data") if isinstance(row, dict) and "data" in row else row
        if not data or not data.get("sku_code"):
            skipped += 1
            continue
        sku = data["sku_code"]
        existing = db.query(Product).filter(Product.sku_code == sku).first()
        if existing:
            if not existing.is_active:
                errors.append(f"{sku}: produto anulado — não atualizado")
                skipped += 1
                continue
            ncm_in = data.get("ncm")
            if ncm_in and existing.ncm and existing.ncm != ncm_in:
                errors.append(f"{sku}: NCM existente não sobrescrito via import")
                data = {k: v for k, v in data.items() if k != "ncm"}
            for field, value in data.items():
                if field == "sku_code":
                    continue
                if value is not None or field in ("weight_kg", "volume_m3", "ncm"):
                    setattr(existing, field, value)
            updated += 1
        else:
            product = Product(**{k: v for k, v in data.items() if v is not None or k in ("weight_kg", "volume_m3")})
            if product.weight_kg is None and "weight_kg" in data:
                product.weight_kg = data["weight_kg"]
            if product.volume_m3 is None and "volume_m3" in data:
                product.volume_m3 = data["volume_m3"]
            db.add(product)
            created += 1

    db.commit()
    return {"created": created, "updated": updated, "skipped": skipped, "errors": errors}


def export_products_xlsx(items: list[dict]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Produtos"
    header = [
        "sku_code", "description", "product_group", "product_subgroup", "lifecycle_status",
        "ncm", "supplier_code", "default_supplier_name", "country_of_origin", "unit_of_measure",
        "weight_kg", "volume_m3", "category", "launch_date", "fiscal_description",
    ]
    ws.append(header)
    for row in items:
        ws.append([row.get(h) for h in header])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
