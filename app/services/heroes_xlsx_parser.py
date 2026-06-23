"""Parser de planilhas Heroes Itália (CONTI ITALIA-BRASILE.xlsx)."""

from __future__ import annotations

import hashlib
import io
import re
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Any

from openpyxl import load_workbook

from app.core.enums import HEROES_XLSX_PARSER_VERSION, HeroesSheetType
from app.core.parse import optional_int
from app.core.parse_it import parse_it_date, parse_it_number
from app.services.heroes_merged_cells import apply_merged_cells_to_grid
from app.services.product_category import suggest_product_category

ORDINE_SHEET_RE = re.compile(r"ordine\s+(\d+)", re.IGNORECASE)
ORDINE_PK_RE = re.compile(r"ordine\s+(\d+)\s+pk", re.IGNORECASE)


def _cell_str(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _norm_header(val: Any) -> str:
    if val is None:
        return ""
    return re.sub(r"\s+", " ", str(val).lower().strip())


def excel_date_to_iso(val: Any) -> str | None:
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, (int, float)):
        base = datetime(1899, 12, 30)
        return (base + timedelta(days=float(val))).date().isoformat()
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def detect_sheet_type(sheet_name: str, grid: list[list[Any]] | None = None) -> str:
    name = sheet_name.strip()
    upper = name.upper()
    if ORDINE_SHEET_RE.search(name):
        return HeroesSheetType.ORDER.value
    if "NON GRAFICATE" in upper:
        return HeroesSheetType.FUTURE_PLANNING.value
    if re.fullmatch(r"20\d{2}", name.strip()):
        year = name.strip()
        if grid:
            flat = " ".join(_norm_header(c) for row in grid[:15] for c in row if c is not None)
            if "data valuta" in flat and "numero fattura" in flat:
                return HeroesSheetType.FINANCIAL_ANNUAL.value
            if year in ("2025", "2026"):
                return HeroesSheetType.FINANCIAL_ANNUAL.value
            return HeroesSheetType.FUTURE_PLANNING.value
        return HeroesSheetType.FINANCIAL_ANNUAL.value if year in ("2025", "2026") else HeroesSheetType.FUTURE_PLANNING.value
    if "RITIRI" in upper or "HK" in upper:
        return HeroesSheetType.LOGISTICS.value
    if "RACCHETTE DA RICEVERE" in upper or "DA RICEVERE" in upper:
        return HeroesSheetType.RECEIPT_AGGREGATE.value
    if grid:
        flat = " ".join(_norm_header(c) for row in grid[:15] for c in row if c is not None)
        if "data valuta" in flat and "numero fattura" in flat:
            return HeroesSheetType.FINANCIAL_ANNUAL.value
        if "da spedire" in flat and ("fattura" in flat or "quantit" in flat):
            return HeroesSheetType.ORDER.value
        if "invoice" in flat and "prodotto" in flat:
            return HeroesSheetType.LOGISTICS.value
    return HeroesSheetType.UNKNOWN.value


def list_xlsx_sheets(content: bytes) -> list[dict[str, Any]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    out = []
    for name in wb.sheetnames:
        ws = wb[name]
        sample: list[list[Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 20:
                break
            sample.append(list(row))
        sheet_type = detect_sheet_type(name, sample)
        order_info = _find_order_numbers(sample if sample else [], name)
        order_match = ORDINE_SHEET_RE.search(name)
        out.append(
            {
                "sheet_name": name,
                "sheet_type": sheet_type,
                "order_number_hint": order_match.group(1) if order_match else order_info.get("order_number_from_sheet_name"),
                "order_number_from_content": order_info.get("order_number_from_content"),
                "order_number_divergence": order_info.get("order_number_divergence", False),
            }
        )
    wb.close()
    return out


def _grid_from_sheet(ws) -> tuple[list[list[Any]], int]:
    rows: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        if any(c is not None and str(c).strip() for c in row):
            rows.append(list(row))
    merge_count = apply_merged_cells_to_grid(rows, ws)
    return rows, merge_count


def _order_from_sheet_name(sheet_name: str) -> str | None:
    m = ORDINE_PK_RE.search(sheet_name) or ORDINE_SHEET_RE.search(sheet_name)
    return m.group(1) if m else None


def _order_from_sheet_content(grid: list[list[Any]]) -> str | None:
    for row in grid[:25]:
        for cell in row:
            s = _cell_str(cell)
            if s:
                m2 = ORDINE_SHEET_RE.search(s)
                if m2:
                    return m2.group(1)
    return None


def _parse_versato(grid: list[list[Any]]) -> dict[str, Any] | None:
    """Extrai versato do topo da sheet — informação legada, não pagamento."""
    for ri, row in enumerate(grid[:10]):
        for ci, cell in enumerate(row[:12]):
            label = _norm_header(cell)
            if label == "versato" or label.startswith("versato "):
                for cj in range(ci + 1, min(ci + 5, len(row))):
                    num = parse_it_number(row[cj])
                    if num is not None and num > 0:
                        col_letter = chr(ord("A") + cj) if cj < 26 else f"col{cj}"
                        return {
                            "versato_amount": str(num),
                            "versato_currency": "EUR",
                            "versato_source_row": ri + 1,
                            "versato_source_cell": f"{col_letter}{ri + 1}",
                            "versato_raw_value": _cell_str(row[cj]),
                            "versato_confidence": 0.9,
                        }
                if ri + 1 < len(grid):
                    below = grid[ri + 1]
                    if ci < len(below):
                        num = parse_it_number(below[ci])
                        if num is not None and num > 0:
                            col_letter = chr(ord("A") + ci) if ci < 26 else f"col{ci}"
                            return {
                                "versato_amount": str(num),
                                "versato_currency": "EUR",
                                "versato_source_row": ri + 2,
                                "versato_source_cell": f"{col_letter}{ri + 2}",
                                "versato_raw_value": _cell_str(below[ci]),
                                "versato_confidence": 0.85,
                            }
    return None


def _find_order_numbers(grid: list[list[Any]], sheet_name: str) -> dict[str, Any]:
    from_name = _order_from_sheet_name(sheet_name)
    from_content = _order_from_sheet_content(grid)
    confirmed = from_content or from_name
    divergence = bool(from_name and from_content and from_name != from_content)
    return {
        "order_number_from_sheet_name": from_name,
        "order_number_from_content": from_content,
        "order_number": confirmed,
        "order_number_divergence": divergence,
    }


def _find_invoice_header_row(grid: list[list[Any]]) -> tuple[int, dict[str, int], str | None] | None:
    for ri, row in enumerate(grid):
        headers = {_norm_header(c): ci for ci, c in enumerate(row) if c is not None}
        keys = set(headers.keys())
        has_data = any("data" == k or k.startswith("data") for k in keys)
        has_fattura = any("fattura" in k for k in keys)
        has_qty = any("quantit" in k for k in keys)
        if has_data and has_fattura and has_qty:
            col_map: dict[str, int] = {}
            product_header: str | None = None
            for k, ci in headers.items():
                if k == "data" or k.startswith("data"):
                    col_map["date"] = ci
                elif "fattura" in k:
                    col_map["invoice_number"] = ci
                elif "quantit" in k:
                    col_map["quantity"] = ci
                elif ("racchetta" in k or "articolo" in k or k == "prodotto") and "credito" not in k:
                    col_map["product"] = ci
                    product_header = "articolo" if "articolo" in k else "racchetta"
                elif "acconto rimasto" in k or "acconto rimanente" in k:
                    col_map["acconto_remaining"] = ci
                elif k == "acconto" or (k.startswith("acconto") and "rimast" not in k and "riman" not in k):
                    col_map["acconto"] = ci
                elif "credito" in k and ("racchetta" in k or "unit" in k or "/" in k):
                    col_map["credit_per_unit"] = ci
                elif "credito accumulato" in k or "credito accum" in k:
                    col_map["credit_accumulated"] = ci
            if "product" not in col_map:
                for k, ci in headers.items():
                    if "racchetta" in k or "articolo" in k:
                        col_map["product"] = ci
                        product_header = "articolo" if "articolo" in k else "racchetta"
                        break
            return ri, col_map, product_header
    return None


def _get_col(row: list[Any], col_map: dict[str, int], key: str) -> Any:
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_invoice_block(grid: list[list[Any]], sheet_name: str) -> tuple[list[dict], list[str], list[str]]:
    warnings: list[str] = []
    errors: list[str] = []
    header = _find_invoice_header_row(grid)
    if not header:
        errors.append("Cabeçalho de faturas não encontrado (data + fattura + quantità)")
        return [], warnings, errors

    header_row, col_map, default_product_header = header
    da_spedire_row = None
    for ri, row in enumerate(grid):
        for cell in row:
            if cell and "DA SPEDIRE" in str(cell).upper():
                da_spedire_row = ri
                break

    items: list[dict] = []
    current_date: str | None = None
    current_invoice: str | None = None

    for ri in range(header_row + 1, da_spedire_row or len(grid)):
        row = grid[ri]
        if not any(c is not None and str(c).strip() for c in row):
            continue
        row_text = " ".join(_norm_header(c) for c in row if c)
        if "da spedire" in row_text:
            break

        raw_date = _get_col(row, col_map, "date")
        raw_inv = _get_col(row, col_map, "invoice_number")
        if raw_date is not None and str(raw_date).strip():
            iso, date_review = parse_it_date(raw_date)
            current_date = iso
            if date_review:
                warnings.append(f"Linha {ri + 1}: data ambígua — revisar")
        if raw_inv is not None and str(raw_inv).strip():
            current_invoice = _cell_str(raw_inv)

        product_raw = _cell_str(_get_col(row, col_map, "product"))
        qty = optional_int(_get_col(row, col_map, "quantity"))
        acconto = parse_it_number(_get_col(row, col_map, "acconto"))
        acconto_rem = parse_it_number(_get_col(row, col_map, "acconto_remaining"))
        credit_unit = parse_it_number(_get_col(row, col_map, "credit_per_unit"))
        credit_acc = parse_it_number(_get_col(row, col_map, "credit_accumulated"))

        if not product_raw and qty is None and acconto is None:
            continue
        if not product_raw:
            warnings.append(f"Linha {ri + 1}: produto vazio — ignorada")
            continue

        cat, conf, rev = suggest_product_category(product_raw)
        product_header = default_product_header or "racchetta"
        needs_review = rev or conf < 0.6 or not current_invoice
        item = {
            "row_number": ri + 1,
            "invoice_date": current_date,
            "invoice_number": current_invoice,
            "item_quantity": qty,
            "product_name_raw": product_raw,
            "acconto_amount": str(acconto) if acconto is not None else None,
            "acconto_remaining": str(acconto_rem) if acconto_rem is not None else None,
            "credit_per_unit": str(credit_unit) if credit_unit is not None else None,
            "credit_accumulated": str(credit_acc) if credit_acc is not None else None,
            "suggested_category": cat,
            "category_confidence": conf,
            "category_review": rev,
            "source_column_product_header": product_header,
            "parser_confidence": 0.9 if current_invoice else 0.7,
            "needs_review": needs_review,
            "raw_values": [str(c) if c is not None else "" for c in row],
        }
        if not current_invoice:
            warnings.append(f"Linha {ri + 1}: fatura herdada/ausente — revisar")
        items.append(item)

    return items, warnings, errors


def _parse_da_spedire(grid: list[list[Any]]) -> tuple[list[dict], list[str]]:
    warnings: list[str] = []
    start = None
    for ri, row in enumerate(grid):
        for cell in row:
            if cell and "DA SPEDIRE" in str(cell).upper():
                start = ri
                break
        if start is not None:
            break

    if start is None:
        return [], ["Bloco DA SPEDIRE não encontrado"]

    header_row = None
    col_map: dict[str, int] = {}
    for ri in range(start, min(start + 5, len(grid))):
        row = grid[ri]
        headers = {_norm_header(c): ci for ci, c in enumerate(row) if c is not None}
        has_product_header = any("prodotto" in k or "racchetta" in k or "articolo" in k for k in headers)
        has_price_header = any("listino" in k for k in headers) or any(
            "fattura" in k and "prezzo" in k for k in headers
        )
        if not has_product_header and not has_price_header:
            continue
        header_row = ri
        for k, ci in headers.items():
            if "racchetta" in k or "articolo" in k or "prodotto" in k:
                col_map["product"] = ci
            elif "quantit" in k or "q.tà" in k:
                col_map["quantity"] = ci
            elif "listino" in k:
                col_map["price_listino"] = ci
            elif "fattura" in k and "prezzo" in k:
                col_map["invoice_price"] = ci
            elif "sconto" in k or "discount" in k:
                col_map["discount"] = ci
            elif "acconto" in k and "riman" not in k and "credito" not in k:
                col_map["acconto"] = ci
            elif "credito" in k and "riman" in k:
                col_map["credit_remaining"] = ci
            elif "note" in k:
                col_map["notes"] = ci
        break

    if header_row is None:
        return [], ["Cabeçalho DA SPEDIRE não encontrado"]

    if "product" not in col_map:
        for dr in range(header_row + 1, min(header_row + 10, len(grid))):
            data_row = grid[dr]
            for ci, cell in enumerate(data_row):
                s = _cell_str(cell)
                if not s:
                    continue
                h = _norm_header(cell)
                if h in ("da spedire",) or "listino" in h or ("prezzo" in h and "fattura" in h):
                    continue
                if parse_it_number(cell) is not None and not isinstance(cell, str):
                    continue
                col_map["product"] = ci
                break
            if "product" in col_map:
                break

    if "quantity" not in col_map and "product" in col_map:
        pi = col_map["product"]
        used_cols = set(col_map.values())
        for dr in range(header_row + 1, min(header_row + 10, len(grid))):
            data_row = grid[dr]
            for ci, cell in enumerate(data_row):
                if ci == pi or ci in used_cols:
                    continue
                qty = optional_int(cell)
                if qty is not None and qty <= 10_000:
                    col_map["quantity"] = ci
                    break
            if "quantity" in col_map:
                break

    if header_row is None or "product" not in col_map:
        return [], ["Cabeçalho DA SPEDIRE não encontrado"]

    rows_out: list[dict] = []
    for ri in range(header_row + 1, len(grid)):
        row = grid[ri]
        product = _cell_str(_get_col(row, col_map, "product"))
        if not product or _norm_header(product) == "da spedire":
            continue
        listino_raw = _get_col(row, col_map, "price_listino")
        if isinstance(listino_raw, str) and "listino" in _norm_header(listino_raw):
            continue
        qty = optional_int(_get_col(row, col_map, "quantity"))
        cat, conf, rev = suggest_product_category(product)
        rows_out.append(
            {
                "row_number": ri + 1,
                "product_name_raw": product,
                "quantity_to_dispatch": qty,
                "price_listino": str(v) if (v := parse_it_number(_get_col(row, col_map, "price_listino"))) is not None else None,
                "invoice_price": str(v) if (v := parse_it_number(_get_col(row, col_map, "invoice_price"))) is not None else None,
                "discount": str(v) if (v := parse_it_number(_get_col(row, col_map, "discount"))) is not None else None,
                "acconto": str(v) if (v := parse_it_number(_get_col(row, col_map, "acconto"))) is not None else None,
                "credit_remaining": str(v) if (v := parse_it_number(_get_col(row, col_map, "credit_remaining"))) is not None else None,
                "notes": _cell_str(_get_col(row, col_map, "notes")),
                "suggested_category": cat,
                "category_confidence": conf,
                "category_review": rev,
                "raw_values": [str(c) if c is not None else "" for c in row],
            }
        )

    deduped: dict[str, dict] = {}
    for row in rows_out:
        key = row["product_name_raw"].lower().strip()
        prev = deduped.get(key)
        if prev is None:
            deduped[key] = row
            continue
        prev_score = (prev.get("quantity_to_dispatch") or 0, 1 if prev.get("price_listino") else 0)
        row_score = (row.get("quantity_to_dispatch") or 0, 1 if row.get("price_listino") else 0)
        if row_score > prev_score:
            deduped[key] = row
    return list(deduped.values()), warnings


def _parse_financial_sheet(grid: list[list[Any]], sheet_name: str) -> dict[str, Any]:
    payments: list[dict] = []
    warnings = ["Sheet financeira — preview auxiliar; pagamentos ambíguos não serão importados automaticamente"]
    for ri, row in enumerate(grid):
        headers_hint = _norm_header(row[0]) if row else ""
        if "data valuta" in " ".join(_norm_header(c) for c in row):
            continue
        if len(row) >= 4 and row[0] is not None:
            desc = _cell_str(row[2] if len(row) > 2 else None)
            if desc and ("fattura" in desc.lower() or "accredit" in desc.lower()):
                payments.append(
                    {
                        "row_number": ri + 1,
                        "date": excel_date_to_iso(row[0]),
                        "invoice_number": _cell_str(row[1] if len(row) > 1 else None),
                        "description": desc,
                        "amount": str(v) if (v := parse_it_number(row[3] if len(row) > 3 else None)) is not None else None,
                        "linked_order": None,
                        "confidence": 0.5,
                    }
                )
    return {"sheet_name": sheet_name, "payments_preview": payments, "warnings": warnings}


def _parse_logistics_sheet(grid: list[list[Any]], sheet_name: str) -> dict[str, Any]:
    header_row = None
    col_map: dict[str, int] = {}
    for ri, row in enumerate(grid[:30]):
        headers = {_norm_header(c): ci for ci, c in enumerate(row) if c is not None}
        if "invoice" in headers and ("prodotto" in headers or "product" in headers):
            header_row = ri
            for k, ci in headers.items():
                if k == "invoice":
                    col_map["invoice"] = ci
                elif "prodotto" in k:
                    col_map["product"] = ci
                elif k == "po":
                    col_map["po"] = ci
                elif "sea" in k or "air" in k:
                    col_map["modal"] = ci
                elif "box" in k and "n" in k:
                    col_map["boxes"] = ci
                elif "pcs per box" in k or "pcs/box" in k:
                    col_map["pcs_per_box"] = ci
                elif "tot pcs" in k or "totale" in k:
                    col_map["total_pcs"] = ci
                elif "stato" in k or "status" in k:
                    col_map["status"] = ci
            break

    lines: list[dict] = []
    if header_row is not None:
        for ri in range(header_row + 1, len(grid)):
            row = grid[ri]
            inv = _cell_str(_get_col(row, col_map, "invoice"))
            prod = _cell_str(_get_col(row, col_map, "product"))
            if not inv and not prod:
                continue
            lines.append(
                {
                    "row_number": ri + 1,
                    "invoice_number": inv,
                    "product_name_raw": prod,
                    "po_reference": _cell_str(_get_col(row, col_map, "po")),
                    "modal": _cell_str(_get_col(row, col_map, "modal")),
                    "boxes": optional_int(_get_col(row, col_map, "boxes")),
                    "pcs_per_box": optional_int(_get_col(row, col_map, "pcs_per_box")),
                    "total_pcs": optional_int(_get_col(row, col_map, "total_pcs")),
                    "status": _cell_str(_get_col(row, col_map, "status")),
                }
            )
    return {"sheet_name": sheet_name, "logistics_lines": lines}


def parse_xlsx_sheet(content: bytes, sheet_name: str, *, file_checksum: str | None = None) -> dict[str, Any]:
    wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' não encontrada")
    ws = wb[sheet_name]
    grid, merged_count = _grid_from_sheet(ws)
    wb.close()

    sheet_type = detect_sheet_type(sheet_name, grid)
    order_info = _find_order_numbers(grid, sheet_name)
    order_number = order_info["order_number"]
    warnings: list[str] = []
    errors: list[str] = []

    if order_info["order_number_divergence"]:
        warnings.append(
            f"Divergência ordem: nome da sheet sugere {order_info['order_number_from_sheet_name']}, "
            f"conteúdo indica {order_info['order_number_from_content']} — confirmar antes do commit"
        )

    result: dict[str, Any] = {
        "format_version": "Heroes Order Import Format v1",
        "parser_version": HEROES_XLSX_PARSER_VERSION,
        "source_file": None,
        "file_checksum": file_checksum,
        "sheet_name": sheet_name,
        "sheet_type": sheet_type,
        "order_number": order_number,
        "order_number_from_sheet_name": order_info["order_number_from_sheet_name"],
        "order_number_from_content": order_info["order_number_from_content"],
        "order_number_divergence": order_info["order_number_divergence"],
        "confirmed_order_number": None,
        "legacy_sheet_summary": None,
        "supplier": "Heroes Itália",
        "currency": "EUR",
        "merged_cell_count": merged_count,
        "locale_numeric": "it-IT",
        "invoice_items": [],
        "da_spedire": [],
        "financial_preview": None,
        "logistics_preview": None,
        "aggregate_preview": None,
        "warnings": warnings,
        "errors": errors,
        "ignored_rows": [],
    }

    if sheet_type == HeroesSheetType.ORDER.value:
        result["legacy_sheet_summary"] = _parse_versato(grid)
        items, w, e = _parse_invoice_block(grid, sheet_name)
        result["invoice_items"] = items
        warnings.extend(w)
        errors.extend(e)
        da_rows, da_w = _parse_da_spedire(grid)
        result["da_spedire"] = da_rows
        warnings.extend(da_w)
        # Produtos únicos
        names = {i["product_name_raw"] for i in items} | {d["product_name_raw"] for d in da_rows}
        result["new_products"] = []
        for n in sorted(names):
            cat, conf, rev = suggest_product_category(n)
            result["new_products"].append(
                {
                    "product_name_raw": n,
                    "suggested_category": cat,
                    "category_confidence": conf,
                    "category_review": rev,
                }
            )
        # Faturas agrupadas
        inv_map: dict[str, dict] = {}
        for it in items:
            inv = it.get("invoice_number") or "—"
            if inv not in inv_map:
                inv_map[inv] = {"invoice_number": inv, "invoice_date": it.get("invoice_date"), "items": []}
            inv_map[inv]["items"].append(it)
        result["invoices_detected"] = list(inv_map.values())

    elif sheet_type == HeroesSheetType.FINANCIAL_ANNUAL.value:
        result["financial_preview"] = _parse_financial_sheet(grid, sheet_name)

    elif sheet_type == HeroesSheetType.LOGISTICS.value:
        result["logistics_preview"] = _parse_logistics_sheet(grid, sheet_name)

    elif sheet_type == HeroesSheetType.RECEIPT_AGGREGATE.value:
        result["aggregate_preview"] = {"note": "Agregação de recebimento — QA/conferência apenas", "row_count": len(grid)}
        warnings.append("Sheet agregada — não usada como fonte oficial primária")

    elif sheet_type == HeroesSheetType.FUTURE_PLANNING.value:
        result["aggregate_preview"] = {"note": "Planejamento futuro — QA apenas", "row_count": len(grid)}
        warnings.append("Sheet de planejamento — não importar automaticamente")

    else:
        warnings.append(f"Tipo de sheet desconhecido: {sheet_name}")

    result["warnings"] = warnings
    result["errors"] = errors
    return result


def make_idempotency_key(file_checksum: str, sheet_name: str) -> str:
    raw = f"{file_checksum}|{sheet_name}|{HEROES_XLSX_PARSER_VERSION}"
    return hashlib.sha256(raw.encode()).hexdigest()
