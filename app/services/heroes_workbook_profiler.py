"""Profiling read-only da planilha legada Heroes (sem escrita no banco)."""

from __future__ import annotations

from decimal import Decimal

import hashlib
import io
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.core.enums import HEROES_XLSX_PARSER_VERSION, HeroesSheetType
from app.core.parse_it import detect_currencies_in_values, format_parse_example, parse_it_number
from app.services.heroes_merged_cells import count_merged_cells
from app.services.heroes_workbook_paths import HEROES_WORKBOOK_FILENAME, resolve_heroes_workbook_path
from app.services.heroes_xlsx_parser import (
    ORDINE_PK_RE,
    ORDINE_SHEET_RE,
    _cell_str,
    _find_invoice_header_row,
    _norm_header,
    detect_sheet_type,
)

RECOMMENDATIONS = {
    "import": "importar",
    "auxiliary": "usar como auxiliar",
    "qa": "usar só para QA",
    "ignore": "ignorar",
    "manual": "revisão manual",
}

TOTAL_TOLERANCE = 0.02


def _file_checksum(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def _order_from_name(sheet_name: str) -> str | None:
    m = ORDINE_PK_RE.search(sheet_name) or ORDINE_SHEET_RE.search(sheet_name)
    return m.group(1) if m else None


def _order_from_content(grid: list[list[Any]], limit_rows: int = 30) -> str | None:
    for row in grid[:limit_rows]:
        for cell in row:
            s = _cell_str(cell)
            if s:
                m = ORDINE_SHEET_RE.search(s)
                if m:
                    return m.group(1)
    return None


def _sheet_used_range(ws) -> dict[str, int | str | None]:
    min_row = ws.min_row or 1
    max_row = ws.max_row or 0
    min_col = ws.min_column or 1
    max_col = ws.max_column or 0
    return {
        "min_row": min_row,
        "max_row": max_row,
        "min_col": min_col,
        "max_col": max_col,
        "range": f"{min_row}:{max_row} x {min_col}:{max_col}" if max_row else None,
        "row_count": max(0, max_row - min_row + 1) if max_row else 0,
        "column_count": max(0, max_col - min_col + 1) if max_col else 0,
    }


def _detect_headers_and_fields(grid: list[list[Any]]) -> tuple[list[str], list[str], list[str]]:
    headers: set[str] = set()
    recognized: set[str] = set()
    for row in grid[:40]:
        for cell in row:
            h = _norm_header(cell)
            if not h:
                continue
            headers.add(h)
            if any(k in h for k in ("data", "fattura", "quantit", "racchetta", "articolo", "acconto", "credito")):
                recognized.add(h)
            if "da spedire" in h:
                recognized.add("da spedire")
            if h in ("invoice", "prodotto", "po", "stato"):
                recognized.add(h)

    expected_order = {"data", "fattura", "quantità", "quantita", "racchetta", "articolo", "da spedire"}
    missing = sorted(k for k in expected_order if not any(k in h for h in headers))
    return sorted(headers), sorted(recognized), missing


def _classify_with_content(sheet_name: str, grid: list[list[Any]]) -> str:
    base = detect_sheet_type(sheet_name, grid)
    upper = sheet_name.strip().upper()
    flat = " ".join(_norm_header(c) for row in grid[:25] for c in row if c is not None)

    if "NON GRAFICATE" in upper or ("2027" in sheet_name and "ordine" not in flat.lower() and "da spedire" not in flat):
        if base == HeroesSheetType.FINANCIAL_ANNUAL.value and "data valuta" in flat:
            return HeroesSheetType.FINANCIAL_ANNUAL.value
        if "non graficate" in upper or ("2027" in sheet_name and "plan" in flat):
            return HeroesSheetType.FUTURE_PLANNING.value
        if sheet_name.strip() == "2027" and "data valuta" not in flat:
            return HeroesSheetType.FUTURE_PLANNING.value

    if base == HeroesSheetType.UNKNOWN.value and "FLOKY" in upper:
        return HeroesSheetType.UNKNOWN.value

    return base


def _parser_confidence(
    sheet_type: str,
    order_name: str | None,
    order_content: str | None,
    grid: list[list[Any]],
    has_da_spedire: bool,
) -> float:
    score = 0.35
    if sheet_type == HeroesSheetType.ORDER.value:
        score += 0.25
        if _find_invoice_header_row(grid):
            score += 0.15
        if has_da_spedire:
            score += 0.1
        if order_name and order_content and order_name == order_content:
            score += 0.15
        elif order_name and order_content and order_name != order_content:
            score -= 0.25
    elif sheet_type in (
        HeroesSheetType.LOGISTICS.value,
        HeroesSheetType.FINANCIAL_ANNUAL.value,
        HeroesSheetType.RECEIPT_AGGREGATE.value,
    ):
        score += 0.35 if grid else 0.1
    elif sheet_type == HeroesSheetType.FUTURE_PLANNING.value:
        score += 0.2
    return max(0.0, min(1.0, score))


def _recommendation(sheet_type: str, confidence: float, divergence: bool, errors: list[str]) -> str:
    if divergence or errors:
        return RECOMMENDATIONS["manual"]
    if sheet_type == HeroesSheetType.ORDER.value:
        if confidence >= 0.75:
            return RECOMMENDATIONS["import"]
        if confidence >= 0.5:
            return RECOMMENDATIONS["manual"]
        return RECOMMENDATIONS["ignore"]
    if sheet_type == HeroesSheetType.FINANCIAL_ANNUAL.value:
        return RECOMMENDATIONS["auxiliary"]
    if sheet_type == HeroesSheetType.LOGISTICS.value:
        return RECOMMENDATIONS["auxiliary"]
    if sheet_type == HeroesSheetType.RECEIPT_AGGREGATE.value:
        return RECOMMENDATIONS["qa"]
    if sheet_type == HeroesSheetType.FUTURE_PLANNING.value:
        return RECOMMENDATIONS["qa"]
    if sheet_type == HeroesSheetType.UNKNOWN.value:
        return RECOMMENDATIONS["manual"]
    return RECOMMENDATIONS["ignore"]


def _collect_numeric_samples(grid: list[list[Any]], limit: int = 8) -> list[dict[str, str | None]]:
    samples: list[dict[str, str | None]] = []
    for row in grid:
        for cell in row:
            if cell is None:
                continue
            s = str(cell)
            if re.search(r"\d", s) and ("," in s or "€" in s or "R$" in s or "." in s):
                parsed = parse_it_number(cell)
                samples.append(format_parse_example(cell, parsed))
                if len(samples) >= limit:
                    return samples
    return samples


def _find_declared_total(grid: list[list[Any]]) -> tuple[str | None, Decimal | None]:
    for row in grid[:15]:
        row_text = " ".join(_norm_header(c) for c in row if c)
        if "versato" in row_text or "totale" in row_text:
            for cell in row:
                parsed = parse_it_number(cell)
                if parsed is not None and parsed > 0:
                    return _cell_str(cell), parsed
    return None, None


def _sum_parsed_amounts(grid: list[list[Any]]) -> Decimal | None:
    header = _find_invoice_header_row(grid)
    if not header:
        return None
    header_row, col_map, _ = header
    acconto_idx = col_map.get("acconto")
    if acconto_idx is None:
        return None
    total = Decimal("0")
    found = False
    for row in grid[header_row + 1 :]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        if any("da spedire" in _norm_header(c) for c in row if c):
            break
        val = parse_it_number(row[acconto_idx] if acconto_idx < len(row) else None)
        if val is not None:
            total += val
            found = True
    return total if found else None


def profile_sheet_from_grid(sheet_name: str, grid: list[list[Any]], *, merged_count: int = 0) -> dict[str, Any]:
    sheet_type = _classify_with_content(sheet_name, grid)
    order_name = _order_from_name(sheet_name)
    order_content = _order_from_content(grid)
    divergence = bool(order_name and order_content and order_name != order_content)
    headers, recognized, missing = _detect_headers_and_fields(grid)
    flat = " ".join(_norm_header(c) for row in grid[:30] for c in row if c)
    has_da_spedire = "da spedire" in flat

    confidence = _parser_confidence(sheet_type, order_name, order_content, grid, has_da_spedire)
    errors: list[str] = []
    warnings: list[str] = []
    if divergence:
        warnings.append(
            f"Divergência: sheet '{sheet_name}' sugere ordine {order_name}, conteúdo indica ordine {order_content}"
        )
        confidence = max(0.0, confidence - 0.3)

    currencies = detect_currencies_in_values([c for row in grid for c in row if c is not None])
    if "EUR" in currencies and "BRL" in currencies:
        warnings.append("Moedas EUR e BRL detectadas na mesma sheet — não somar totais mistos")

    declared_raw, declared_total = _find_declared_total(grid)
    summed = _sum_parsed_amounts(grid)
    total_check: dict[str, Any] | None = None
    if declared_total is not None and summed is not None:
        diff = abs(declared_total - summed)
        total_check = {
            "declared_raw": declared_raw,
            "declared_total": str(declared_total),
            "parsed_sum": str(summed),
            "difference": str(diff),
            "needs_review": diff > TOTAL_TOLERANCE,
        }
        if diff > TOTAL_TOLERANCE:
            warnings.append("Possível linha perdida ou merge mal lido — total declarado ≠ soma parseada")

    recommendation = _recommendation(sheet_type, confidence, divergence, errors)

    return {
        "sheet_name": sheet_name,
        "used_range": None,
        "row_count": len(grid),
        "column_count": max((len(r) for r in grid), default=0),
        "sheet_type": sheet_type,
        "order_number_from_sheet_name": order_name,
        "order_number_from_content": order_content,
        "order_number_divergence": divergence,
        "detected_headers": headers[:40],
        "recognized_fields": recognized,
        "missing_expected_fields": missing,
        "has_da_spedire_block": has_da_spedire,
        "merged_cell_count": merged_count,
        "locale_numeric": "it-IT",
        "numeric_parse_examples": _collect_numeric_samples(grid),
        "currencies_detected": currencies,
        "total_validation": total_check,
        "parser_confidence": round(confidence, 3),
        "recommendation": recommendation,
        "warnings": warnings,
        "errors": errors,
    }


def profile_workbook_bytes(content: bytes, *, source_path: str | None = None) -> dict[str, Any]:
    checksum = _file_checksum(content)
    wb_ro = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_names = list(wb_ro.sheetnames)

    sheets_profile: list[dict[str, Any]] = []
    for name in sheet_names:
        ws = wb_ro[name]
        grid: list[list[Any]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= 200:
                break
            grid.append(list(row))
        used = _sheet_used_range(ws)
        profile = profile_sheet_from_grid(name, grid, merged_count=0)
        profile["used_range"] = used.get("range")
        profile["row_count"] = used.get("row_count") or profile["row_count"]
        profile["column_count"] = used.get("column_count") or profile["column_count"]
        sheets_profile.append(profile)
    wb_ro.close()

    wb_merge = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    merge_by_sheet = {n: count_merged_cells(wb_merge[n]) for n in sheet_names}
    wb_merge.close()
    for sp in sheets_profile:
        sp["merged_cell_count"] = merge_by_sheet.get(sp["sheet_name"], 0)

    return {
        "profiler_version": HEROES_XLSX_PARSER_VERSION,
        "source_file": source_path or HEROES_WORKBOOK_FILENAME,
        "file_checksum": checksum,
        "sheet_count": len(sheet_names),
        "sheets": sheets_profile,
        "database_writes": False,
        "read_only_mode": True,
        "note": "Profiling é fisicamente incapaz de gravar no banco — apenas leitura do XLSX.",
    }


def profile_workbook_file(path: Path | str | None = None) -> dict[str, Any]:
    resolved = resolve_heroes_workbook_path(path)
    if resolved is None:
        searched = [str(p) for p in (__import__("app.services.heroes_workbook_paths", fromlist=["SEARCH_PATHS"]).SEARCH_PATHS)]
        raise FileNotFoundError(
            f"Planilha não encontrada. Procurado em: {', '.join(searched)}"
        )
    content = resolved.read_bytes()
    result = profile_workbook_bytes(content, source_path=str(resolved))
    result["resolved_path"] = str(resolved)
    return result
