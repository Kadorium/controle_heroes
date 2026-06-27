"""Testes build_invoice_blocks e guards L-UX-001-A."""

from __future__ import annotations

import io
import uuid

import pytest
from openpyxl import load_workbook

from app.config import ROOT_DIR
from app.core.enums import HeroesImportRunStatus
from app.models import HeroesImportRun, Invoice, Payment
from app.services.heroes_invoice_blocks import build_invoice_blocks, get_invoice_blocks_for_preview
from app.services.heroes_order_format_v1 import preview_to_canonical
from app.services.heroes_workbook_paths import HEROES_WORKBOOK_FILENAME
from app.services.heroes_xlsx_commit import commit_heroes_import_run
from app.services.heroes_xlsx_import import preview_xlsx_sheet, register_workbook_bytes
from app.services.heroes_xlsx_parser import parse_xlsx_sheet
from tests.fixtures.heroes_xlsx_builder import build_ordine_758_xlsx

REAL_WORKBOOK = ROOT_DIR / HEROES_WORKBOOK_FILENAME


def _unique_758_xlsx() -> bytes:
    buf = build_ordine_758_xlsx()
    wb = load_workbook(io.BytesIO(buf))
    ws = wb.active
    ws["Z1"] = uuid.uuid4().hex
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def test_build_invoice_blocks_strips_acconto_from_items():
    flat = [
        {
            "row_number": 1,
            "invoice_number": "F-100",
            "invoice_date": "2025-03-10",
            "product_name_raw": "STARLIGHT",
            "item_quantity": 10,
            "acconto_amount": "500",
        },
        {
            "row_number": 2,
            "invoice_number": "F-100",
            "invoice_date": "2025-03-10",
            "product_name_raw": "AURA",
            "item_quantity": 5,
            "acconto_amount": None,
        },
    ]
    blocks = build_invoice_blocks(flat)
    assert len(blocks) == 1
    assert len(blocks[0]["acconto_payments"]) == 1
    assert blocks[0]["acconto_payments"][0]["receipt_reference"] == "ACCONTO-F-100"
    assert "acconto_amount" not in blocks[0]["items"][0]
    assert len(blocks[0]["items"]) == 2


def test_build_invoice_blocks_dedupes_same_amount_merge():
    flat = [
        {"row_number": 1, "invoice_number": "181.0", "product_name_raw": "a", "acconto_amount": "74000"},
        {"row_number": 2, "invoice_number": "181.0", "product_name_raw": "b", "acconto_amount": "74000"},
        {"row_number": 3, "invoice_number": "181.0", "product_name_raw": "c", "acconto_amount": "74000"},
    ]
    blocks = build_invoice_blocks(flat)
    assert len(blocks[0]["acconto_payments"]) == 1
    assert blocks[0]["acconto_payments"][0]["amount"] == "74000"


def test_build_invoice_blocks_multiple_independent_acconto_per_invoice():
    flat = [
        {"row_number": 9, "invoice_number": "6.0", "product_name_raw": "aura", "acconto_amount": "5000"},
        {"row_number": 10, "invoice_number": "6.0", "product_name_raw": "fierce", "acconto_amount": "10000"},
        {"row_number": 11, "invoice_number": "6.0", "product_name_raw": "bull", "acconto_amount": "10000"},
    ]
    blocks = build_invoice_blocks(flat)
    payments = blocks[0]["acconto_payments"]
    assert len(payments) == 3
    refs = {p["receipt_reference"] for p in payments}
    assert refs == {"ACCONTO-6.0-9", "ACCONTO-6.0-10", "ACCONTO-6.0-11"}


@pytest.mark.skipif(not REAL_WORKBOOK.is_file(), reason="Planilha real ausente")
def test_real_ordine_758_invoice_blocks_181_and_737():
    preview = parse_xlsx_sheet(REAL_WORKBOOK.read_bytes(), "Ordine 758")
    blocks = get_invoice_blocks_for_preview(preview)
    b181 = next(b for b in blocks if str(b.get("invoice_number")) == "181.0")
    assert len(b181["acconto_payments"]) == 1
    assert b181["acconto_payments"][0]["amount"] == "74000"
    assert len(b181["items"]) == 5
    assert all("acconto_amount" not in it for it in b181["items"])

    b737 = next(b for b in blocks if str(b.get("invoice_number")) == "737.0")
    assert len(b737["acconto_payments"]) == 1
    assert b737["acconto_payments"][0]["amount"] == "10000"


@pytest.mark.skipif(not REAL_WORKBOOK.is_file(), reason="Planilha real ausente")
def test_real_ordine_758_invoice_6_three_payments():
    preview = parse_xlsx_sheet(REAL_WORKBOOK.read_bytes(), "Ordine 758")
    b6 = next(b for b in preview["invoice_blocks"] if str(b.get("invoice_number")) == "6.0")
    assert len(b6["acconto_payments"]) == 3


@pytest.mark.skipif(not REAL_WORKBOOK.is_file(), reason="Planilha real ausente")
def test_real_commit_payment_counts(db):
    order_number = f"758-a-{uuid.uuid4().hex[:8]}"
    content = REAL_WORKBOOK.read_bytes()
    reg = register_workbook_bytes(db, content, filename="ord758.xlsx", user_id=1)
    run = preview_xlsx_sheet(
        db,
        raw_file_id=reg["raw_file_id"],
        sheet_name="Ordine 758",
        content=content,
        filename="ord758.xlsx",
        user_id=1,
        confirmed_order_number=order_number,
    )
    imp = commit_heroes_import_run(
        db,
        run.id,
        user_id=1,
        confirm_import=True,
        confirm_sheet_match=True,
        confirmed_order_number=order_number,
    )
    inv181 = db.query(Invoice).filter(
        Invoice.importation_id == imp.id, Invoice.invoice_number == "181.0"
    ).first()
    assert inv181 is not None
    pays181 = db.query(Payment).filter(Payment.invoice_id == inv181.id).all()
    assert len(pays181) == 1
    assert pays181[0].amount_foreign == 74000

    inv737 = db.query(Invoice).filter(
        Invoice.importation_id == imp.id, Invoice.invoice_number == "737.0"
    ).first()
    assert inv737 is not None
    pays737 = db.query(Payment).filter(Payment.invoice_id == inv737.id).all()
    assert len(pays737) == 1

    inv6 = db.query(Invoice).filter(
        Invoice.importation_id == imp.id, Invoice.invoice_number == "6.0"
    ).first()
    assert inv6 is not None
    pays6 = db.query(Payment).filter(Payment.invoice_id == inv6.id).all()
    assert len(pays6) == 3


@pytest.mark.skipif(not REAL_WORKBOOK.is_file(), reason="Planilha real ausente")
def test_preview_to_canonical_paid_total_from_block_not_line_sum():
    preview = parse_xlsx_sheet(REAL_WORKBOOK.read_bytes(), "Ordine 758")
    canonical = preview_to_canonical(preview)
    inv181 = next(i for i in canonical["invoices"] if i["invoice_number"] == "181.0")
    assert inv181["paid_total"] == "74000.0" or inv181["paid_total"] == "74000"
    inv6 = next(i for i in canonical["invoices"] if i["invoice_number"] == "6.0")
    assert float(inv6["paid_total"]) == 25000.0


def test_attached_raw_blocks_standalone_preview(admin_client, db):
    uid = uuid.uuid4().hex[:8]
    sup = admin_client.post(
        "/api/suppliers",
        json={"name": f"Guard {uid}", "country": "IT", "currency_default": "EUR"},
    ).json()
    imp = admin_client.post(
        "/api/importations",
        json={
            "po_number": f"GUARD-{uid}",
            "supplier_id": sup["id"],
            "currency": "EUR",
            "incoterm": "FOB",
        },
    ).json()
    content = _unique_758_xlsx()
    upload = admin_client.post(
        "/api/imports/heroes/xlsx/upload",
        files={
            "file": (
                "ord758.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    raw_id = upload.json()["raw_file_id"]

    prev_before = admin_client.post(
        "/api/imports/heroes/xlsx/preview",
        json={"raw_file_id": raw_id, "sheet_name": "Ordine 758"},
    )
    assert prev_before.status_code == 200
    standalone_run_id = prev_before.json()["run_id"]

    link = admin_client.post(
        f"/api/importations/{imp['id']}/link-heroes-raw",
        json={"raw_file_id": raw_id},
    )
    assert link.status_code == 201

    prev = admin_client.post(
        "/api/imports/heroes/xlsx/preview",
        json={"raw_file_id": raw_id, "sheet_name": "Ordine 758"},
    )
    assert prev.status_code == 409
    assert "Central da ordem" in prev.json()["detail"]

    commit = admin_client.post(
        "/api/imports/heroes/xlsx/commit",
        json={
            "run_id": standalone_run_id,
            "confirm_import": True,
            "confirm_sheet_match": True,
            "confirmed_order_number": f"758-{uid}",
        },
    )
    assert commit.status_code == 409

    attached = db.query(HeroesImportRun).filter(
        HeroesImportRun.id == link.json()["run_id"]
    ).first()
    assert attached.status == HeroesImportRunStatus.ATTACHED.value
