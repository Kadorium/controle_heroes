"""Gera XLSX de teste simulando Ordine 758 da planilha Heroes."""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook


def build_ordine_758_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ordine 758"

    ws["A1"] = "versato"
    ws["A2"] = "ordine 758"
    headers = [
        "data",
        "n* fattura",
        "quantità",
        "racchetta",
        "acconto",
        "acconto rimasto",
        "credito/ racchetta",
    ]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)

    rows = [
        (datetime(2025, 3, 10), "F-100", 10, "STARLIGHT 300", 500, 1500, None),
        (None, None, 5, "AURA POWER", None, None, 2.5),
        (datetime(2025, 4, 1), "F-101", 20, "palline beach", 200, None, None),
        (datetime(2025, 4, 15), "F-102", 3, "WASHBAG EPIC", None, None, None),
    ]
    for ri, row in enumerate(rows, 5):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    da_row = 5 + len(rows) + 1
    ws.cell(row=da_row, column=1, value="DA SPEDIRE")
    da_headers = ["racchetta", "quantità", "prezzo listino", "prezzo fattura", "sconto"]
    for ci, h in enumerate(da_headers, 1):
        ws.cell(row=da_row + 1, column=ci, value=h)
    da_data = [
        ("STARLIGHT 300", 8, 120, 100, 5),
        ("palline beach", 15, 8, 7.5, None),
        ("WASHBAG EPIC", 3, 25, 22, None),
    ]
    for ri, row in enumerate(da_data, da_row + 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    # Sheet logística
    ws2 = wb.create_sheet("RITIRI HK")
    ws2.append(["INVOICE", "PRODOTTO", "PO", "SEA/AIR", "N° BOX", "PCS PER BOX", "TOT PCS", "STATO"])
    ws2.append(["F-100", "STARLIGHT 300", "758", "AIR", 2, 5, 10, "SHIPPED"])
    ws2.append(["F-102", "WASHBAG EPIC", "758", "SEA", 1, 3, 3, "PLANNED"])

    # Sheet financeira
    ws3 = wb.create_sheet("2025")
    ws3.append(["Data Valuta", "Numero fattura", "Descrizione", "Valuta", "Accrediti"])
    ws3.append([datetime(2025, 5, 1), "F-100", "PARZIALE FATTURA SALDATO CON ACCONTO", "EUR", 500])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_ordine_759_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ordine 759"
    ws["A2"] = "ordine 759"
    headers = ["data", "n* fattura", "quantità", "articolo", "acconto", "acconto rimasto"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=4, column=ci, value=h)
    ws.append([datetime(2025, 2, 1), "F-200", 6, "OLYMPIA", 100, 400])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
